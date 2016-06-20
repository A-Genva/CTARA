#CTARA Project
import  itertools #Package for additional functions
import math #Package for maths functions
import numpy as np
import openpyxl as pyxl
from openpyxl import Workbook
from decimal import *

getcontext().prec = 4

points = list() #List to store values of points entered by the user
score = list() #List to store score of individual points entered by the user
Xp = list() #List to store the x-coordinate of a point entered by the user
Yp = list() #List to store the y-coordinate of a point entered by the user

wb = pyxl.load_workbook('Excel.xlsx')
ws = wb.active
i=0
for column in ws.columns:
    for cell in (column):
        if i==0:
            points.append(cell.value)
        elif i==1:
            score.append(cell.value)
        elif i==2:
            Xp.append(cell.value)
        elif i==3:
            Yp.append(cell.value)
        
    i+=1
 
Nc = (int(input("Input number of clusters: "))) #Input from user for no. of clusters
Nppc = (int(input("Input number of points per cluster: ")))
"""
def distance(lat1, lon1, lat2, lon2): #Function to calculate the distance b/w latitude and longitude of two points
    p = 0.017453292519943295
    a = 0.5 - cos((lat2 - lat1) * p)/2 + cos(lat1 * p) * cos(lat2 * p) * (1 - cos((lon2 - lon1) * p)) / 2
    return 12742 * asin(sqrt(a))
"""
def dist(Xa,Ya,Xb,Yb): #Defining a function to calculate the distance b/w two points
    d = math.sqrt(((Xa-Xb)**2)+((Ya-Yb)**2))
    return d

c1=list() #List to store the values of cluster 1
c2=list() #List to store the values of cluster 2
c3=list() #List to store the values of cluster 3

l = list(itertools.combinations(points,Nppc)) #Function to determine the possible combinations of all the points
for i in range(int(len(l)/Nppc)):
    l1 = list(itertools.combinations(list(set(points).difference(l[i])),Nppc)) #Using concept of sets to evaluate the combinations of remaining set
    for j in range(int(len(l1))):
        c1.append(l[i])
        if set(l1[j]).intersection(set(l[i]))==set(): #Checking for disjoint sets
            c2.append(l1[j])
            c3.append(set(points).difference(set(l[i]).union(set(l1[j])))) #Checking for disjoint sets
    

def sumOfscores(x): #Function to determine the sum of scores of all points in a cluster
    temp=0
    for i in range(len(x)):
        for j in range(len(points)):
            if x[i]==points[j]: #Comparing each point in a cluster with the points list
                temp+=int(score[j])
    return temp

sc1 = list() #List to store the values of sum of scores of all points in a cluster for cluster 1
sc2 = list() #List to store the values of sum of scores of all points in a cluster for cluster 2
sc3 = list() #List to store the values of sum of scores of all points in a cluster for cluster 3

for i in range(len(c1)):
    sc1.append(sumOfscores(list(c1[i]))) #Storing the values of sum of scores of all points in a cluster for cluster 1
    sc2.append(sumOfscores(list(c2[i]))) #Storing the values of sum of scores of all points in a cluster for cluster 2
    sc3.append(sumOfscores(list(c3[i]))) #Storing the values of sum of scores of all points in a cluster for cluster 3
    
def normalized_list(X): #Function to determine the normalized list of a particular list
    nl = list()
    for i in range(len(X)):
        nl.append(X[i]/max(X)) #Dividing each element of a list by the maximum value element of that list
    return nl

def coeffVar(X,Y,Z): #Function to determine the coefficient of variation for 3 points
        mean = 0.0
        stdev = 0.0
        mean = (X+Y+Z)/2.0 #Calculating mean for 3 points
        stdev = math.sqrt((X-mean)**2.0+(Y-mean)**2.0+(Z-mean)**2) #Calculating standard deviation of 3 points
        return  stdev/mean

cv1 = list() #List to store the values of coefficient of variation for the scores of each cluster

for i in range(len(sc1)):
    cv1.append(coeffVar(float(sc1[i]),float(sc2[i]),float(sc3[i]))) #Storing values of coefficient of variation for the scores of each cluster
    
ncv1 = list() #List to store the normalized values of cv1
ncv1.append(normalized_list(cv1)) #Calling the function to get the normalized list

def d(cluster): #Defining a function to calculate the minimum sum of distances from the node to other points in a cluster 
    val = list() #Dummy list for storing values
    for i in range(len(cluster)):
        for j in range(len(points)):
            if cluster[i]==points[j]:
                val.append(j)    
    a=dist(int(Xp[val[0]]),int(Yp[val[0]]),int(Xp[val[1]]),int(Yp[val[1]])) #Calling the distance function to determine the distance b/w two points in a cluster
    b=dist(int(Xp[val[0]]),int(Yp[val[0]]),int(Xp[val[2]]),int(Yp[val[2]]))
    c=dist(int(Xp[val[2]]),int(Yp[val[2]]),int(Xp[val[1]]),int(Yp[val[1]]))        
    return min(a+b,a+c,b+c) #Returning the minimum sum of value for each cluster

dc1 = list() #List to store the values of minimum sum of distances from the node to other points in a cluster 1
dc2 = list() #List to store the values of minimum sum of distances from the node to other points in a cluster 2
dc3 = list() #List to store the values of minimum sum of distances from the node to other points in a cluster 3

for i in range(len(c1)): #Storing the values of minimum sum of distances from the node to other points for each cluster set
    dc1.append(d(list(c1[i])))
    dc2.append(d(list(c2[i])))
    dc3.append(d(list(c3[i])))

def func(d1,d2,d3): #Defining a function to account the minimum of (d1-d2) , (d2-d3) and (d1-d3)
    return math.sqrt((d1-d2)**2+(d1-d3)**2+(d2-d3)**2)

fgd = list() #List to store the values obtained from the above function

for i in range(len(dc1)): #Storing values in the list
    fgd.append(func(dc1[i],dc2[i],dc3[i]))

nfgd = list() #List to store the normalized values of fgd 
nfgd.append(normalized_list(fgd))

def func1(d1,d2,d3): #Defining a function to obtain the square of least distances in a cluster
    return ((d1**2)+(d2**2)+(d3**2)) #Evaluating the value
    
fds = list() #List to store the values of the square of least distances in a cluster
for i in range(len(dc1)):
    fds.append(func1(dc1[i],dc2[i],dc3[i]))

nfds = list() #List to store the normalized values of fds
nfds.append(normalized_list(fds))  

def global_func(a,b,c): #Defining a global function which takes account of all parameters  
    s = math.sqrt((a**2)+(b**2)+(c**2)) #Distance formula
    return float(s)
    
global_list = list() #List to store the values of global_func
for i in range(len(c1)):
    global_list.append(global_func(float(nfds[0][i]),float(nfgd[0][i]),float(ncv1[0][i])))
    
m = min(global_list)#To find the minimum value in the list final_value
print(m)
index = global_list.index(min(global_list)) #Determining the index of global_list which has a minimum value


wb1 = Workbook(write_only=True)
ws = wb1.create_sheet()
k = len(c1)
j=0
for irow in range(13):            
    if j==0:
        ws.append([str(c1[i]) for i in range(k)])
    if j==1:
        ws.append([str(c2[i]) for i in range(k)])
    if j==2:
        ws.append([str(c3[i]) for i in range(k)])
    if j==3:
        ws.append([cv1[i] for i in range(k)])
    if j==4:
        ws.append([ncv1[0][i] for i in range(k)])
    if j==5:
        ws.append([dc1[i] for i in range(k)])
    if j==6:
        ws.append([dc2[i] for i in range(k)])
    if j==7:
        ws.append([dc3[i] for i in range(k)])
    if j==8:
        ws.append([fgd[i] for i in range(k)])
    if j==9:
        ws.append([nfgd[0][i] for i in range(k)])
    if j==10:
        ws.append([fds[i] for i in range(k)])
    if j==11:
        ws.append([nfds[0][i] for i in range(k)])
    if j==12:
        ws.append([global_list[i] for i in range(k)])
    
    
    j+=1
            
wb1.save('NewExcel.xlsx')

print("The minimum value considering all the parameters is for cluster "+str(c1[index])+" , "+str(c2[index])+" and "+str(c3[index]))























